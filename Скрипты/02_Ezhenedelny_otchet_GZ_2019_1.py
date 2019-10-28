import openpyxl
import os
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
fil = os.path.abspath(__file__)

tab=[
        {
        "indicator": 1,
        "category": 1,
        "unit": 1,
        "work": 1,
        "result": "\u0427\u0442\u043e \u0442\u043e \u043f\u043e\u043b\u0435\u0437\u043d\u043e\u0435",
        "need": 213,
        "accept": 123,
        "date": "2019-10-15"
        },
        {
        "indicator": 1,
        "category": 2,
        "unit": 1,
        "work": 1,
        "result": "2",
        "need": 213,
        "accept": 123,
        "date": "2019-10-16"
        },
        {
        "indicator": 1,
        "category": 2,
        "unit": 1,
        "work": 1,
        "result": "2",
        "need": 213,
        "accept": 123,
        "date": "2019-10-16"
        },
        {
        "indicator": 2,
        "category": 1,
        "unit": 1,
        "work": 1,
        "result": "3",
        "need": 213,
        "accept": 123,
        "date": "2019-10-16"
        },
        {
        "indicator": 2,
        "category": 1,
        "unit": 1,
        "work": 1,
        "result": "4",
        "need": 213,
        "accept": 123,
        "date": "2019-10-16"
        },
        {
        "indicator": 3,
        "category": 1,
        "unit": 1,
        "work": 1,
        "result": "5",
        "need": 213,
        "accept": 123,
        "date": "2019-10-16"
        },
        {
        "indicator": 3,
        "category": 1,
        "unit": 1,
        "work": 1,
        "result": "6",
        "need": 213,
        "accept": 123,
        "date": "2019-10-16"
        },
    ]

wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1,column=2).value='Отчет о выполнении государственного задания в 2019 году'
ws.merge_cells(start_row=1,start_column=2,end_row=1,end_column=5)
ws.cell(row=1,column=6).value='III квартал'
ws.merge_cells(start_row=1,start_column=6,end_row=1,end_column=8)

rows=['Наименование показателя','Категория','Ответственное подразделение','Виды работ','Результат','Необходимо','Выполнено','Процент']
for i, value in enumerate(rows,1):
    ws.cell(row=2,column= i).value=value

sh=0
sc=0


value_ind=0
value_categ=0
mas_indicator=[]
mas_indicator_col=[]
mas_category=[]
mas_category_col=[]
w=-1
for i in tab:
    if i["indicator"]!=value_ind:
        value_ind=i["indicator"]
        mas_indicator.append(value_ind)
        mas_indicator_col.append(1)
        w+=1
    else:
        mas_indicator_col[w]+=1


print(mas_indicator) 
print(mas_indicator_col) 



w=-1
f=-1
for i in mas_indicator:
    mas_category.append([])
    mas_category_col.append([])
    f+=1
    value_categ=0
    for y in tab:
       
        if y.get("indicator")==i:

            if y.get("category")!= value_categ:
                value_categ=y.get("category")
                mas_category[f].append(value_categ)
                mas_category_col[f].append(1)
                
                w+=1
            else:
                mas_category_col[f][w]+=1
    w=-1

print(mas_category) 
print(mas_category_col) 


e=0
q=0
a=0
aa=0
for i, rows in enumerate(tab,1):
    
    if i==mas_indicator_col[e]+1+q:
        e+=1
        q=mas_indicator_col[e]
        a+=1
        aa=a
    else:
        aa=a
    print(i,e,q,a,aa)

    for y, row in enumerate(list(rows.keys()),1):

        cell=ws.cell(row=i+aa+2, column=y)
        cell.value=rows[row]
    

f=0
g=0
for i in mas_indicator_col:

    ws.merge_cells(start_row=3+f,start_column=1,end_row=3+f+i-1,end_column=1)

    for y in mas_category_col:
        for z in y:
            ws.merge_cells(start_row=3+g,start_column=2,end_row=3+g+z-1,end_column=2)
            g+=z
        g+=1
    f+=i
    ws.cell(row=3+f,column=1).value='Итого:'
    ws.merge_cells(start_row=3+f,start_column=1,end_row=3+f,end_column=5)
    f+=1
# Форматирование таблицы
thin_border=Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
                            
style1 = Font(name='PT Astra Serif',size=11,bold=True,vertAlign='baseline')
style2 = Font(name='PT Astra Serif',size=11,bold=False,vertAlign='baseline')
for cell in ws['A1':'H2']:
    for row in cell:
        row.font= style1
        row.alignment=Alignment(horizontal='center',vertical='center',wrapText=True)
        if row.row ==2:
            row.border = thin_border 
# как f используется кол-во строк записей        
for cell in ws['A3':f'H{f+2}']:
    for row in cell:
        row.font= style2
        row.alignment=Alignment(horizontal='center',vertical='center',wrapText=True)
        row.border = thin_border 
        

ws.column_dimensions['A'].width=26.71
ws.column_dimensions['B'].width=16.57
ws.column_dimensions['C'].width=17.71
ws.column_dimensions['D'].width=48.57
ws.column_dimensions['E'].width=38.71
ws.column_dimensions['F'].width=14.71
ws.column_dimensions['G'].width=13.71
ws.column_dimensions['H'].width=13.71



fil = os.path.join(fil,'..\..\Документы')
fil = os.path.abspath(fil)
os.chdir(path=fil)
wb.save('02_Ezhenedelny_otchet_GZ_2019_1.xlsx')



